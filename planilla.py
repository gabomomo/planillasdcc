#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Planilla Bisemanal - Ingenieria Estrella S.A.

Toma el archivo de horas (DCC) y llena la Planilla Madre:
  - consolida las filas partidas por jornada
  - separa jornada diurna / mixta / nocturna por el divisor de la hora
  - traduce los COLORES de las celdas de dia a dias de ausencia / incapacidad
  - resuelve cedula -> carne (aplicando la tabla de correcciones de la hoja Errores)
  - escribe la tabla ControlHoras y el roster de Nomina Quincena
  - genera un reporte de revision con todo lo que necesita ojo humano

No modifica los archivos originales: escribe una copia nueva.
"""

import os
import re
import sys
import shutil
import datetime
import unicodedata
from collections import defaultdict

# La libreria de Excel viene incluida en programa/lib para que el programa
# funcione en cualquier computadora sin instalar nada.
_LIB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib")
if os.path.isdir(_LIB) and _LIB not in sys.path:
    sys.path.insert(0, _LIB)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\n  FALTA LA LIBRERIA DE EXCEL.")
    print("  Debe existir la carpeta:  programa/lib/openpyxl")
    print("  Copie de nuevo la carpeta 'programa' completa.\n")
    sys.exit(1)


# ----------------------------------------------------------------------------
# Configuracion de estructura (si Excel cambia, se ajusta aqui)
# ----------------------------------------------------------------------------

# --- Archivo de horas (DCC) -------------------------------------------------
H_HOJA = "Horas de trabajo "
H_NOMINA = "Nomina"
H_FILA_ENC = 2          # fila de encabezados
H_FILA_INI = 3          # primera fila de datos
H_COL_JORNADA = 2       # B  Tipo de jornada (D/M/N)
H_COL_NOMBRE = 3        # C
H_COL_CEDULA = 4        # D
H_COL_INGRESO = 5       # E
H_COL_DEPTO = 6         # F
H_COL_PUESTO = 7        # G
H_COL_DIA1 = 8          # H  primer dia del periodo
H_COL_EXTRAS = 23       # W  horas extra 150%
H_COL_DOBLES = 24       # X  horas dobles sencillas
H_COL_EXTRA_DOB = 25    # Y  horas extra dobles 200%
H_COL_AJ_ORD = 26       # Z  ajuste horas ordinarias
H_COL_AJ_EXT = 27       # AA ajuste horas extra
H_COL_NOVEDAD = 30      # AD notas

N_COL_SALARIO = 12      # L  Salario Contrato
N_COL_HORA = 14         # N  ¢/hora (de aqui se lee el divisor de jornada)
N_COL_SUBTOTAL = 30     # AD subtotal salarios brutos
N_COL_INCENTIVO = 35    # AI incentivos / reintegros
N_COL_BRUTO = 34        # AH salario bruto
N_COL_RENTA = 37        # AK impuesto de renta
N_COL_ADELANTO = 38     # AL adelanto de salario
N_COL_PENSION = 39      # AM pension alimentaria
N_COL_EPP = 40          # AN rebajo por EPP

# Colores de la leyenda (relleno solido de las celdas de dia)
COLORES = {
    "FFFBE2D5": "PSG",
    "FFF8CBAD": "PSG",
    "FFFCE4D6": "PSG",
    "FFFF0000": "AUSENCIA",
    "FF00B0F0": "CCSS",
    "FF00B050": "INS",
    "FF4EA72E": "INS",
    "FF00B04F": "INS",
}

# --- Planilla Madre ---------------------------------------------------------
M_BASE = "Base de datos del Personal"
M_CONTROL = "Control de Horas 2"
M_NOMINA = "Nomina Quincena (1)"
M_ERRORES = "Errores"
M_RENTAS = "Rentas"

BD_FILA_INI = 3
BD_COL_CARNE = 1        # A
BD_COL_CEDULA = 6       # F
BD_COL_INGRESO = 12     # L
BD_COL_SALIDA = 13      # M
BD_COL_SALARIO = 18     # R

CH_FILA_INI = 3
CH_COL = {              # columnas de la tabla ControlHoras
    "carne": 1, "horas_ord": 9, "extras": 10, "dias_dobles": 11, "extra_dobles": 12,
    "horas_mixtas": 13, "mixtas_extra": 14, "horas_noct": 15, "noct_extra": 16,
    "ausencia": 17, "psg": 18, "ccss": 19, "ccss_mas3": 20, "ins": 21,
}
CH_COL_FORMULA = [2, 3, 4, 5, 6, 7, 8]   # B..H se rellenan con la formula plantilla

NQ_FILA_INI = 3
NQ_COL_CARNE = 1
NQ_COL_VACACIONES = 45  # AS
NQ_COL_REINTEGROS = 46  # AT
NQ_COL_ANTICIPOS = 49   # AW
NQ_COL_PENSION = 50     # AX
NQ_COL_RENTA = 51       # AY  retencion del impuesto de renta
# La formula de AY se reconstruye siempre: cuando un pago ABRE ventana se escribe
# un 0 literal, y si se tomara la plantilla de esa corrida la formula se perderia.
# Ademas el rango original estaba fijo en $O$347 y se queda corto si crece el roster.
NQ_RENTA_FORMULA = ("=VLOOKUP(Tabla2[[#This Row],[Carn\u00e9]],"
                    "Rentas!$K$1:$O$%d,5,FALSE)")
NQ_COL_NOTAS = 53       # BA
NQ_ULT_COL = 53         # BA
# columnas de Nomina Quincena que se escriben con valor (no formula)
NQ_COL_VALOR = {NQ_COL_VACACIONES, NQ_COL_REINTEGROS, NQ_COL_ANTICIPOS,
                NQ_COL_PENSION, NQ_COL_NOTAS}

JORNADA_HORAS = {8: 8, 7: 7, 6: 6}


# ----------------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------------

def norm_cedula(x):
    """Normaliza una cedula para poder cruzarla entre archivos."""
    if x is None:
        return None
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[^0-9A-Za-z]", "", s)
    s = s.lstrip("0")
    return s or None


def sin_tildes(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()


def num(x):
    return x if isinstance(x, (int, float)) and not isinstance(x, bool) else 0


def a_fecha(x):
    if isinstance(x, datetime.datetime):
        return x.date()
    if isinstance(x, datetime.date):
        return x
    return None


def color_celda(celda):
    f = celda.fill
    if f is None or f.fill_type != "solid":
        return None
    fg = f.fgColor
    if fg is None or fg.type != "rgb":
        return None
    rgb = fg.rgb
    if not isinstance(rgb, str):
        return None
    return COLORES.get(rgb.upper())


def divisor_de(formula, jornada_letra):
    """Saca el divisor de jornada (8/7/6) de la formula =L/30/8 de la hoja Nomina."""
    if isinstance(formula, str):
        m = re.search(r"/\s*30\s*/\s*(\d+)", formula)
        if m:
            d = int(m.group(1))
            if d in JORNADA_HORAS:
                return d
    letra = (str(jornada_letra) or "").strip().upper()[:1]
    return {"D": 8, "M": 7, "N": 6}.get(letra, 8)


def reubicar(formula, fila):
    """Adapta una formula plantilla de la fila 3 a otra fila."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    return re.sub(r"(?<![\$A-Za-z0-9_])([A-Z]{1,2})3(?![0-9])",
                  lambda m: "%s%d" % (m.group(1), fila), formula)


DIAS_CICLO = 14          # el ciclo es BISEMANAL: 14 dias exactos, 26 pagos al ano

# Regla de la empresa para el impuesto de renta:
# la base imponible se arma con DOS pagos consecutivos ("ventana"). Cuando un mes
# calendario recibe 3 pagos, el tercero NO entra en ese mes: se corre y pasa a ser
# el primero de la ventana siguiente. Por eso el ano tiene 13 ventanas, no 12.
RENTA_COL_D = 4          # Rentas: "Salario primera Q ..."  (1er pago de la ventana)
RENTA_COL_I = 9          # Rentas: "Salario Segunda Q ..."  (2do pago de la ventana)
RENTA_COL_K = 11         # Rentas: Carne del bloque de la base imponible
MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre"]


def mes_de_la_ventana(fecha_pago, posicion):
    """La ventana se rotula con el mes del pago que la CIERRA."""
    cierre = fecha_pago if posicion == 2 else fecha_pago + datetime.timedelta(days=DIAS_CICLO)
    return "%s %d" % (MESES[cierre.month], cierre.year)


CARPETA_GENERADAS = "Planillas generadas"
PREFIJO_SALIDA = "Planilla Madre LISTA "
PREFIJO_REPORTE = "Reporte de revision "


def carpeta_generadas(base):
    ruta = os.path.join(base, CARPETA_GENERADAS)
    if not os.path.isdir(ruta):
        os.makedirs(ruta)
    return ruta


def listar_generadas(base):
    """Planillas ya generadas, con su fecha de pago leida de adentro del archivo."""
    ruta = os.path.join(base, CARPETA_GENERADAS)
    salida = []
    if not os.path.isdir(ruta):
        return salida
    for n in sorted(os.listdir(ruta)):
        if not n.lower().endswith(".xlsx") or n.startswith("~$"):
            continue
        if "lista" not in sin_tildes(n):
            continue
        full = os.path.join(ruta, n)
        try:
            wb = openpyxl.load_workbook(full, data_only=True, read_only=True)
            ws = wb[M_CONTROL]
            pago = a_fecha(ws["M1"].value)
            fin = a_fecha(ws["J1"].value)
            ini = a_fecha(ws["G1"].value)
            wb.close()
        except Exception:
            continue
        if pago:
            salida.append({"ruta": full, "nombre": n, "pago": pago,
                           "ini": ini, "fin": fin})
    salida.sort(key=lambda x: x["pago"])
    return salida


def elegir_estado(base, ruta_plantilla, fecha_pago=None):
    """De donde se toma el ESTADO (periodo anterior y ventana de renta).

    La plantilla aporta la estructura y la Base de Datos del Personal; el estado
    lo aporta la ULTIMA planilla generada. Se ignoran las generadas cuya fecha de
    pago sea igual o posterior a la de esta corrida, para que volver a correr un
    mismo periodo no adelante la ventana de renta dos veces.
    """
    previas = listar_generadas(base)
    if fecha_pago:
        previas = [x for x in previas if x["pago"] < fecha_pago]
    if previas:
        u = previas[-1]
        return u["ruta"], u["nombre"], False
    return ruta_plantilla, os.path.basename(ruta_plantilla), True


def leer_estado(ruta):
    """Saca del archivo de estado lo unico que se arrastra de un periodo a otro."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[M_CONTROL]
    est = {
        "ini": a_fecha(ws["G1"].value),
        "fin": a_fecha(ws["J1"].value),
        "pago": a_fecha(ws["M1"].value),
        "rentas": {},
        "roster": set(),
        "prim": 0,
        "segu": 0,
    }
    for fila in range(CH_FILA_INI, ws.max_row + 1):
        c = ws.cell(fila, 1).value
        if c:
            est["roster"].add(str(c).strip())
    wr = wb[M_RENTAS]
    for fila in range(2, wr.max_row + 1):
        for col_c, col_v, slot in ((1, RENTA_COL_D, "D"), (6, RENTA_COL_I, "I")):
            c = wr.cell(fila, col_c).value
            if c:
                est["rentas"].setdefault(str(c).strip(), {})[slot] = \
                    wr.cell(fila, col_v).value
        if num(wr.cell(fila, RENTA_COL_D).value):
            est["prim"] += 1
        if num(wr.cell(fila, RENTA_COL_I).value):
            est["segu"] += 1
    est["etiquetas"] = (wr.cell(1, RENTA_COL_D).value, wr.cell(1, RENTA_COL_I).value)
    # si la 2a casilla ya viene llena, la ventana anterior cerro y esta abre otra
    est["posicion"] = 1 if est["segu"] >= max(1, est["prim"] * 0.5) else 2
    wb.close()
    return est


def tramos_de_renta(ws):
    """Lee la tabla de tramos que vive en la hoja Rentas (columnas Q..T)."""
    tramos = []
    for f in range(3, 9):
        inf, sup, monto = (ws.cell(f, 18).value, ws.cell(f, 19).value,
                           ws.cell(f, 20).value)
        tarifa = ws.cell(f, 17).value
        if inf is None and sup is None:
            continue
        tramos.append({"tarifa": num(tarifa), "inf": num(inf), "sup": num(sup),
                       "monto": num(monto)})
    return tramos


def renta_de(base, tramos):
    """Replica la formula IFS de la hoja Rentas."""
    if not tramos or base <= 0:
        return 0.0
    acum = 0.0
    for i, t in enumerate(tramos):
        if t["sup"] and base <= t["sup"]:
            return (base - t["inf"]) * t["tarifa"] + acum if t["tarifa"] else 0.0
        acum += t["monto"]
    ult = tramos[-1]
    return (base - ult["inf"]) * ult["tarifa"] + (acum - ult["monto"])


HOJA_CCSS = "Datos CCSS"

# Palabras con las que se marca una salida en las novedades del archivo de horas.
PALABRAS_SALIDA = ("despido", "renuncia", "renuncio", "renunci",
                   "abandono", "no volvio", "no vino mas", "no regreso")


def escribir_datos_ccss(wb, registros, dias_periodo, fecha_pago):
    """Deja los datos del periodo como VALORES planos, sin formulas.

    La unificacion mensual para la CCSS los lee de aqui. Asi no depende de que
    alguien haya abierto el archivo en Excel para que existan los numeros.
    """
    if HOJA_CCSS in wb.sheetnames:
        del wb[HOJA_CCSS]
    ws = wb.create_sheet(HOJA_CCSS)
    ws.sheet_state = "hidden"

    ws["A1"] = "Datos del periodo para la unificacion mensual. NO EDITAR."
    ws["A2"] = "inicio"
    ws["B2"] = dias_periodo[0]
    ws["C2"] = "fin"
    ws["D2"] = dias_periodo[-1]
    ws["E2"] = "pago"
    ws["F2"] = fecha_pago

    enc = ["Carne", "Identificacion", "Nombre", "Puesto", "Dias Laborados",
           "Total Salario Bruto", "Dias INS", "Fecha de Ingreso", "Novedades",
           "Origen de la fecha de ingreso"]
    for j, t in enumerate(enc, start=1):
        c = ws.cell(4, j)
        c.value = t
        c.font = Font(bold=True, name="Arial")

    fila = 5
    for reg in registros:
        p = reg["persona"]
        bd = reg["bd"]
        dias = (p.horas[8] / 8.0 + p.horas[7] / 7.0 + p.horas[6] / 6.0)
        ws.cell(fila, 1).value = reg["carne"]
        ws.cell(fila, 2).value = p.cedula
        ws.cell(fila, 3).value = p.nombre
        ws.cell(fila, 4).value = p.puesto
        ws.cell(fila, 5).value = round(dias, 4)
        ws.cell(fila, 6).value = round(p.bruto_madre, 2)
        ws.cell(fila, 7).value = round(p.d_ins, 2)
        # La fecha de ingreso manda desde la Base de Datos del Personal, pero
        # ahi suele venir vacia para la gente nueva: entonces se toma la del
        # archivo de horas, que si la trae.
        ing_bd = bd.get("ingreso")
        if ing_bd:
            ing, origen = ing_bd, "Base de Datos"
            if p.ingreso and p.ingreso != ing_bd:
                origen = "Base de Datos (el archivo de horas dice %s)" % \
                    p.ingreso.strftime("%d/%m/%Y")
        elif p.ingreso:
            ing, origen = p.ingreso, "Archivo de horas (falta en la Base de Datos)"
        else:
            ing, origen = None, "No hay fecha en ningun lado"
        ws.cell(fila, 8).value = ing
        ws.cell(fila, 9).value = " / ".join(p.novedades)[:250] if p.novedades else None
        ws.cell(fila, 10).value = origen
        fila += 1
    return ws


def escribir_verificacion(wb, registros, posicion, avisos, estado=None):
    """Hoja de autocomprobacion: lo que calculo la herramienta vs lo que da Excel.

    Sirve para confirmar de un vistazo que el recalculo de Excel produjo lo
    esperado, sin tener que revisar 364 filas a mano.
    """
    n = len(registros)
    f_ini, f_fin = NQ_FILA_INI, NQ_FILA_INI + n - 1

    tramos = tramos_de_renta(wb[M_RENTAS])
    ant = {}
    if estado:
        for c, slots in estado.get("rentas", {}).items():
            ant[c] = num(slots.get("D"))
    else:
        ws_r = wb[M_RENTAS]
        for f in range(2, ws_r.max_row + 1):
            c = ws_r.cell(f, 1).value
            if c:
                ant[str(c).strip()] = num(ws_r.cell(f, RENTA_COL_D).value)

    horas = bruto = ccss = neto = renta_tot = 0.0
    for reg in registros:
        p = reg["persona"]
        b = p.bruto_madre
        r = 0.0
        if posicion == 2:
            r = renta_de(ant.get(reg["carne"], 0.0) + b, tramos)
        horas += p.horas[8]
        bruto += b
        ccss += b * 0.1083
        renta_tot += r
        neto += b - b * 0.1083 - p.anticipos - p.pension - r

    ws = wb.create_sheet("Verificacion")
    ws.sheet_properties.tabColor = "1D7A4C"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30

    ws["A1"] = "VERIFICACION AUTOMATICA"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A2"] = ("Compara lo que calculo la herramienta contra lo que acaba de "
                "calcular Excel. Si todo dice CUADRA, el archivo esta bien.")
    ws["A2"].font = Font(italic=True, size=10, name="Arial")

    enc = ["Concepto", "Calculado por la herramienta", "Calculado por Excel",
           "Diferencia", "Resultado"]
    for j, t in enumerate(enc, start=1):
        c = ws.cell(4, j)
        c.value = t
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", fgColor="FF44546A")

    filas = [
        ("Trabajadores en la planilla", n,
         "=COUNTA('%s'!A%d:A%d)" % (M_NOMINA, f_ini, f_fin), 0),
        ("Horas ordinarias (jornada diurna)", round(horas, 2),
         "=SUM('%s'!I%d:I%d)" % (M_CONTROL, f_ini, f_fin), 0.01),
        ("Total salario bruto", round(bruto, 2),
         "=SUM('%s'!AU%d:AU%d)" % (M_NOMINA, f_ini, f_fin), 1),
        ("Rebajo CCSS 10.83%", round(ccss, 2),
         "=SUM('%s'!AV%d:AV%d)" % (M_NOMINA, f_ini, f_fin), 1),
        ("Retencion de renta", round(renta_tot, 2),
         "=SUM('%s'!AY%d:AY%d)" % (M_NOMINA, f_ini, f_fin), 1),
        ("Total salario neto", round(neto, 2),
         "=SUM('%s'!AZ%d:AZ%d)" % (M_NOMINA, f_ini, f_fin), 1),
    ]
    fila = 5
    for concepto, esperado, formula, tol in filas:
        ws.cell(fila, 1).value = concepto
        ws.cell(fila, 2).value = esperado
        ws.cell(fila, 3).value = formula
        ws.cell(fila, 4).value = "=C%d-B%d" % (fila, fila)
        ws.cell(fila, 5).value = ('=IF(ISBLANK(C{f}),"Excel no ha recalculado",'
                                  'IF(ABS(D{f})<={t},"CUADRA",'
                                  '"REVISAR"))').format(f=fila, t=tol)
        for j in range(1, 6):
            ws.cell(fila, j).font = Font(name="Arial", size=11,
                                         bold=(j == 5))
        for j in (2, 3, 4):
            ws.cell(fila, j).number_format = '#,##0.00'
        fila += 1

    ws.cell(fila + 1, 1).value = "RESULTADO GENERAL"
    ws.cell(fila + 1, 1).font = Font(bold=True, size=12, name="Arial")
    ws.cell(fila + 1, 2).value = ('=IF(COUNTIF(E5:E%d,"CUADRA")=%d,'
                                  '"TODO CUADRA","REVISAR LAS FILAS DE ARRIBA")'
                                  % (fila - 1, len(filas)))
    ws.cell(fila + 1, 2).font = Font(bold=True, size=12, name="Arial")

    ws.cell(fila + 3, 1).value = "Que hacer si dice REVISAR"
    ws.cell(fila + 3, 1).font = Font(bold=True, name="Arial")
    ws.cell(fila + 4, 1).value = (
        "1) Cierre y vuelva a abrir el archivo: Excel recalcula solo al abrirlo.\n"
        "2) Si sigue distinto, la causa mas comun es que se edito a mano alguna "
        "celda de la Nomina o de la Base de Datos despues de generarla.\n"
        "3) Revise el 'Reporte de revision' del mismo periodo.")
    ws.cell(fila + 4, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=fila + 4, start_column=1,
                   end_row=fila + 6, end_column=5)

    avisos.add("verificacion", [
        "Total salario bruto", round(bruto, 2),
        "Se agrego la hoja 'Verificacion' al archivo: al abrirlo en Excel "
        "confirma sola que el recalculo dio estos mismos numeros."])
    return {"trabajadores": n, "bruto": bruto, "ccss": ccss,
            "renta": renta_tot, "neto": neto}


def escribir_rentas(wb, registros, posicion, fecha_pago, avisos, estado=None):
    """Rueda la hoja Rentas segun la regla del corrimiento.

    Pago que ABRE la ventana  -> 1a casilla = bruto de este pago, 2a vacia.
    Pago que la CIERRA        -> 1a casilla se conserva, 2a = bruto de este pago.
    """
    ws = wb[M_RENTAS]
    ult = ws.max_row

    # Las casillas acumuladas vienen del archivo de ESTADO (la planilla generada
    # el periodo anterior), no de la plantilla, que nunca cambia.
    anterior = dict(estado.get("rentas", {})) if estado else {}
    if not anterior:
        for f in range(2, ult + 1):
            for col_c, col_v, slot in ((1, RENTA_COL_D, "D"), (6, RENTA_COL_I, "I")):
                c = ws.cell(f, col_c).value
                if c:
                    anterior.setdefault(str(c).strip(), {})[slot] = ws.cell(f, col_v).value

    plantilla, plantilla_fila = None, None
    for f in range(2, ult + 1):
        o = ws.cell(f, 15).value
        t = getattr(o, "text", o)
        if isinstance(t, str) and t.startswith("="):
            plantilla, plantilla_fila = t, f
            break

    for f in range(2, ult + 1):
        for c in range(1, 16):
            ws.cell(f, c).value = None

    mes = mes_de_la_ventana(fecha_pago, posicion)
    ws.cell(1, RENTA_COL_D).value = "Salario primer pago %s" % mes
    ws.cell(1, RENTA_COL_I).value = "Salario segundo pago %s" % mes

    en_roster = set()
    for i, reg in enumerate(registros):
        f = 2 + i
        car, p = reg["carne"], reg["persona"]
        en_roster.add(car)
        prev = anterior.get(car, {})
        if posicion == 1:
            v_d, v_i = p.bruto_madre, None
        else:
            v_d = prev.get("D")
            v_i = p.bruto_madre

        ws.cell(f, 1).value = car
        ws.cell(f, 2).value = p.nombre
        ws.cell(f, 3).value = p.cedula
        ws.cell(f, RENTA_COL_D).value = round(v_d, 2) if isinstance(v_d, (int, float)) else 0

        if v_i is not None:
            ws.cell(f, 6).value = car
            ws.cell(f, 7).value = p.nombre
            ws.cell(f, 8).value = p.cedula
            ws.cell(f, RENTA_COL_I).value = round(v_i, 2)

        ws.cell(f, RENTA_COL_K).value = car
        ws.cell(f, 12).value = p.nombre
        ws.cell(f, 13).value = p.cedula
        ws.cell(f, 14).value = "=+D%d+I%d" % (f, f)
        if plantilla:
            ws.cell(f, 15).value = re.sub(
                r"(?<![\$A-Za-z0-9])N%d(?![0-9])" % plantilla_fila, "N%d" % f, plantilla)

    for car, slots in anterior.items():
        if car in en_roster:
            continue
        d = slots.get("D")
        if isinstance(d, (int, float)) and d:
            avisos.add("renta", [car, "Quedo fuera del roster",
                                 round(d, 2),
                                 "Tenia %.2f acumulado en la 1a casilla y no vino en "
                                 "este periodo. Su ventana quedo incompleta." % d])



def pagos_del_mes(fecha_pago):
    """Fechas de pago que caen en el mismo mes calendario, con cadencia de 14 dias."""
    d = fecha_pago
    while (d - datetime.timedelta(days=DIAS_CICLO)).month == fecha_pago.month \
            and (d - datetime.timedelta(days=DIAS_CICLO)).year == fecha_pago.year:
        d -= datetime.timedelta(days=DIAS_CICLO)
    salida = []
    while d.month == fecha_pago.month and d.year == fecha_pago.year:
        salida.append(d)
        d += datetime.timedelta(days=DIAS_CICLO)
    return salida


class Aviso(object):
    """Acumula los hallazgos que van al reporte de revision."""

    def __init__(self):
        self.items = defaultdict(list)

    def add(self, seccion, fila):
        self.items[seccion].append(fila)

    def total(self, seccion):
        return len(self.items.get(seccion, []))


# ----------------------------------------------------------------------------
# Lectura del archivo de horas (DCC)
# ----------------------------------------------------------------------------

class Persona(object):
    def __init__(self, cedula):
        self.cedula = cedula
        self.nombre = None
        self.ingreso = None
        self.depto = None
        self.puesto = None
        self.filas = []
        self.novedades = []
        self.enlaces_rotos = []
        self.horas = {8: 0.0, 7: 0.0, 6: 0.0}
        self.extras = {8: 0.0, 7: 0.0, 6: 0.0}
        self.dias_dobles = 0.0
        self.extra_dobles = 0.0
        self.dias = []            # [{'horas':, 'div':, 'color':}]
        self.salario_dcc = None
        self.bruto_dcc = 0.0
        self.bruto_madre = 0.0
        self.incentivos = 0.0
        self.ajuste_monto = 0.0
        self.anticipos = 0.0
        self.pension = 0.0
        self.renta_dcc = 0.0
        # resultado de la clasificacion
        self.d_ausencia = 0.0
        self.d_psg = 0.0
        self.d_ccss = 0.0
        self.d_ccss_mas3 = 0.0
        self.d_ins = 0.0
        self.d_trabajados = 0.0
        self.horas_mixtas_orig = 0.0
        self.horas_noct_orig = 0.0
        self.d_sin_justificar = []       # dias completos sin horas y sin color
        self.d_parciales = []            # dias con horas incompletas
        self.d_previos = 0.0


def leer_horas(ruta, dias_periodo):
    wbf = openpyxl.load_workbook(ruta, data_only=False)
    wbv = openpyxl.load_workbook(ruta, data_only=True)
    hf, hv = wbf[H_HOJA], wbv[H_HOJA]
    nf, nv = wbf[H_NOMINA], wbv[H_NOMINA]
    ndias = len(dias_periodo)

    personas = {}
    orden = []
    fila = H_FILA_INI
    while fila <= hv.max_row:
        ced = norm_cedula(hv.cell(fila, H_COL_CEDULA).value)
        nombre = hv.cell(fila, H_COL_NOMBRE).value
        if not ced or not nombre:
            fila += 1
            continue

        p = personas.get(ced)
        if p is None:
            p = Persona(ced)
            personas[ced] = p
            orden.append(ced)
        p.filas.append(fila)
        p.nombre = (str(nombre).strip() or p.nombre)
        p.ingreso = a_fecha(hv.cell(fila, H_COL_INGRESO).value) or p.ingreso
        p.depto = hv.cell(fila, H_COL_DEPTO).value or p.depto
        p.puesto = hv.cell(fila, H_COL_PUESTO).value or p.puesto
        nov = hv.cell(fila, H_COL_NOVEDAD).value
        if nov and str(nov).strip():
            p.novedades.append(str(nov).strip())

        div = divisor_de(nf.cell(fila, N_COL_HORA).value,
                         hv.cell(fila, H_COL_JORNADA).value)

        if not p.dias:
            p.dias = [{"horas": 0.0, "div": None, "color": None} for _ in range(ndias)]

        for i in range(ndias):
            col = H_COL_DIA1 + i
            hrs = num(hv.cell(fila, col).value)
            col_c = color_celda(hf.cell(fila, col))
            d = p.dias[i]
            if hrs:
                d["horas"] += hrs
                d["div"] = div
            if col_c and not d["color"]:
                d["color"] = col_c
            if d["div"] is None:
                d["div"] = div

        p.extras[div] += num(hv.cell(fila, H_COL_EXTRAS).value)
        p.dias_dobles += num(hv.cell(fila, H_COL_DOBLES).value) / float(div)
        p.extra_dobles += num(hv.cell(fila, H_COL_EXTRA_DOB).value)

        # ajustes de periodos anteriores -> se pagan como reintegro en colones
        tarifa = num(nv.cell(fila, N_COL_HORA).value)
        p.ajuste_monto += num(hv.cell(fila, H_COL_AJ_ORD).value) * tarifa
        p.ajuste_monto += num(hv.cell(fila, H_COL_AJ_EXT).value) * tarifa * 1.5

        # el archivo de horas se copia a si mismo hacia su hoja Nomina;
        # si falta un enlace, ese dato NO se pago. Se detecta aqui.
        for col_h, col_n, que in ((H_COL_EXTRAS, 16, "horas extra"),
                                  (H_COL_DOBLES, 19, "horas dobles"),
                                  (H_COL_EXTRA_DOB, 22, "horas extra dobles")):
            if num(hv.cell(fila, col_h).value) and nf.cell(fila, col_n).value is None:
                p.enlaces_rotos.append("fila %d: %s sin enlace en la hoja Nomina"
                                       % (fila, que))

        sal = nv.cell(fila, N_COL_SALARIO).value
        if isinstance(sal, (int, float)) and sal:
            p.salario_dcc = max(p.salario_dcc or 0, sal)
        p.bruto_dcc += num(nv.cell(fila, N_COL_BRUTO).value)
        p.incentivos += num(nv.cell(fila, N_COL_INCENTIVO).value)
        p.anticipos += num(nv.cell(fila, N_COL_ADELANTO).value) + num(nv.cell(fila, N_COL_EPP).value)
        p.pension += num(nv.cell(fila, N_COL_PENSION).value)
        p.renta_dcc += num(nv.cell(fila, N_COL_RENTA).value)

        # horas ordinarias reales de esta fila
        hsum = sum(num(hv.cell(fila, H_COL_DIA1 + i).value) for i in range(ndias))
        p.horas[div] += hsum

        fila += 1

    return [personas[c] for c in orden]


def clasificar_dias(p, dias_periodo, dias_blanco_como_ausencia, prorratear_ingreso):
    """Traduce el calendario de colores a dias de ausencia / incapacidad."""
    for i, fecha in enumerate(dias_periodo):
        d = p.dias[i] if i < len(p.dias) else {"horas": 0.0, "div": 8, "color": None}
        jor = JORNADA_HORAS.get(d["div"] or 8, 8)
        hrs = d["horas"]
        col = d["color"]
        frac_no_trabajada = max(0.0, (jor - hrs) / float(jor))
        p.d_trabajados += min(1.0, hrs / float(jor))

        if col == "AUSENCIA":
            p.d_ausencia += 1.0 if hrs == 0 else frac_no_trabajada
        elif col == "PSG":
            p.d_psg += 1.0 if hrs == 0 else frac_no_trabajada
        elif col == "CCSS":
            # media jornada digitada => primeros 3 dias (patrono paga 50%)
            # jornada en blanco  => incapacidad larga (patrono no paga)
            if hrs > 0:
                p.d_ccss += 1.0
            else:
                p.d_ccss_mas3 += 1.0
        elif col == "INS":
            p.d_ins += 1.0 if hrs == 0 else frac_no_trabajada
        else:
            if frac_no_trabajada <= 0:
                continue
            antes_de_ingresar = p.ingreso is not None and fecha < p.ingreso
            if antes_de_ingresar:
                if prorratear_ingreso:
                    p.d_previos += frac_no_trabajada
                    p.d_ausencia += frac_no_trabajada
                continue
            if dias_blanco_como_ausencia:
                p.d_ausencia += frac_no_trabajada
            if hrs > 0:
                p.d_parciales.append((fecha, round(frac_no_trabajada, 4)))
            else:
                p.d_sin_justificar.append((fecha, round(frac_no_trabajada, 4)))


# ----------------------------------------------------------------------------
# Escritura en la Planilla Madre
# ----------------------------------------------------------------------------

def leer_correcciones(wb):
    """Lee la tabla de cedulas mal digitadas de la hoja Errores."""
    corr = {}
    if M_ERRORES not in wb.sheetnames:
        return corr
    ws = wb[M_ERRORES]
    for fila in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if sin_tildes(ws.cell(fila, col).value) == "cedula":
                c_mal, c_bien = col, col + 1
                if sin_tildes(ws.cell(fila, c_bien).value) not in ("correccion",):
                    continue
                f = fila + 1
                while f <= ws.max_row:
                    mal = norm_cedula(ws.cell(f, c_mal).value)
                    bien = norm_cedula(ws.cell(f, c_bien).value)
                    if not mal:
                        break
                    if bien:
                        corr[mal] = bien
                    f += 1
    return corr


def indexar_personal(wb):
    ws = wb[M_BASE]
    por_cedula, por_carne = {}, {}
    for fila in range(BD_FILA_INI, ws.max_row + 1):
        carne = ws.cell(fila, BD_COL_CARNE).value
        if not carne:
            continue
        reg = {
            "carne": str(carne).strip(),
            "cedula": norm_cedula(ws.cell(fila, BD_COL_CEDULA).value),
            "ingreso": a_fecha(ws.cell(fila, BD_COL_INGRESO).value),
            "salida": ws.cell(fila, BD_COL_SALIDA).value,
            "salario": ws.cell(fila, BD_COL_SALARIO).value,
            "fila": fila,
        }
        por_carne[reg["carne"]] = reg
        if reg["cedula"]:
            por_cedula.setdefault(reg["cedula"], reg)
    return por_cedula, por_carne


def limpiar_rango(ws, fila_ini, fila_fin, col_fin):
    for fila in range(fila_ini, fila_fin + 1):
        for col in range(1, col_fin + 1):
            ws.cell(fila, col).value = None


def redimensionar(ws, nombre, filas_datos):
    """Ajusta el rango de la tabla de Excel al numero de filas usado."""
    tabla = ws.tables.get(nombre)
    if tabla is None:
        return None
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", tabla.ref)
    if not m:
        return tabla.ref
    c1, f1, c2, _ = m.groups()
    tot = 1 if (tabla.totalsRowCount or 0) else 0
    nueva = int(f1) + filas_datos + tot
    tabla.ref = "%s%s:%s%d" % (c1, f1, c2, nueva)
    return tabla.ref


def escribir_madre(ruta_madre, salida, registros, dias_periodo, fecha_pago,
                   avisos, posicion=2, estado=None):
    shutil.copyfile(ruta_madre, salida)
    wb = openpyxl.load_workbook(salida)

    ini, fin = dias_periodo[0], dias_periodo[-1]

    # --- fechas del periodo -------------------------------------------------
    for hoja in (M_CONTROL, M_NOMINA):
        ws = wb[hoja]
        ws["G1"] = ini
        ws["J1"] = fin
        ws["M1"] = fecha_pago

    # --- ControlHoras -------------------------------------------------------
    ws = wb[M_CONTROL]
    tabla = ws.tables.get("ControlHoras")
    m = re.match(r"[A-Z]+\d+:([A-Z]+)(\d+)", tabla.ref)
    ult_col_ch = openpyxl.utils.column_index_from_string(m.group(1))
    fin_actual = int(m.group(2))
    tot_ch = 1 if (tabla.totalsRowCount or 0) else 0
    fila_totales = None
    if tot_ch:
        fila_totales = {c: ws.cell(fin_actual, c).value for c in range(1, ult_col_ch + 1)}

    plantilla_ch = {c: ws.cell(CH_FILA_INI, c).value for c in CH_COL_FORMULA}
    limpiar_rango(ws, CH_FILA_INI, fin_actual, ult_col_ch)

    for i, reg in enumerate(registros):
        fila = CH_FILA_INI + i
        p = reg["persona"]
        ws.cell(fila, CH_COL["carne"]).value = reg["carne"]
        for c in CH_COL_FORMULA:
            ws.cell(fila, c).value = reubicar(plantilla_ch[c], fila)
        ws.cell(fila, CH_COL["horas_ord"]).value = round(p.horas[8], 4)
        ws.cell(fila, CH_COL["extras"]).value = round(p.extras[8], 4)
        ws.cell(fila, CH_COL["dias_dobles"]).value = round(p.dias_dobles, 4)
        ws.cell(fila, CH_COL["extra_dobles"]).value = round(p.extra_dobles, 4)
        ws.cell(fila, CH_COL["horas_mixtas"]).value = round(p.horas[7], 4)
        ws.cell(fila, CH_COL["mixtas_extra"]).value = round(p.extras[7], 4)
        ws.cell(fila, CH_COL["horas_noct"]).value = round(p.horas[6], 4)
        ws.cell(fila, CH_COL["noct_extra"]).value = round(p.extras[6], 4)
        ws.cell(fila, CH_COL["ausencia"]).value = round(p.d_ausencia, 4)
        ws.cell(fila, CH_COL["psg"]).value = round(p.d_psg, 4)
        ws.cell(fila, CH_COL["ccss"]).value = round(p.d_ccss, 4)
        ws.cell(fila, CH_COL["ccss_mas3"]).value = round(p.d_ccss_mas3, 4)
        ws.cell(fila, CH_COL["ins"]).value = round(p.d_ins, 4)

    if tot_ch:
        fila_t = CH_FILA_INI + len(registros)
        for c, v in fila_totales.items():
            ws.cell(fila_t, c).value = v
    redimensionar(ws, "ControlHoras", len(registros))

    # --- Nomina Quincena ----------------------------------------------------
    ws = wb[M_NOMINA]
    tabla = ws.tables.get("Tabla2")
    m = re.match(r"[A-Z]+\d+:([A-Z]+)(\d+)", tabla.ref)
    ult_col_nq = openpyxl.utils.column_index_from_string(m.group(1))
    fin_actual = int(m.group(2))
    tot_nq = 1 if (tabla.totalsRowCount or 0) else 0
    fila_totales = None
    if tot_nq:
        fila_totales = {c: ws.cell(fin_actual, c).value for c in range(1, ult_col_nq + 1)}

    plantilla_nq = {}
    for c in range(2, ult_col_nq + 1):
        v = ws.cell(NQ_FILA_INI, c).value
        if isinstance(v, str) and v.startswith("="):
            plantilla_nq[c] = v
    limpiar_rango(ws, NQ_FILA_INI, fin_actual, ult_col_nq)

    for i, reg in enumerate(registros):
        fila = NQ_FILA_INI + i
        p = reg["persona"]
        ws.cell(fila, NQ_COL_CARNE).value = reg["carne"]
        for c, f in plantilla_nq.items():
            if c in NQ_COL_VALOR:
                continue
            if c == NQ_COL_RENTA:
                continue
            ws.cell(fila, c).value = reubicar(f, fila)
        # AY siempre se reconstruye, nunca se hereda de la corrida anterior
        ws.cell(fila, NQ_COL_RENTA).value = (
            0 if posicion == 1 else NQ_RENTA_FORMULA % (1 + len(registros)))
        reintegro = round(p.incentivos + p.ajuste_monto, 2)
        ws.cell(fila, NQ_COL_REINTEGROS).value = reintegro if reintegro else 0
        ws.cell(fila, NQ_COL_ANTICIPOS).value = round(p.anticipos, 2)
        ws.cell(fila, NQ_COL_PENSION).value = round(p.pension, 2)
        if p.novedades:
            ws.cell(fila, NQ_COL_NOTAS).value = " / ".join(p.novedades)[:250]

    if tot_nq:
        fila_t = NQ_FILA_INI + len(registros)
        for c, v in fila_totales.items():
            ws.cell(fila_t, c).value = v
    redimensionar(ws, "Tabla2", len(registros))

    # --- Rentas: se rueda la ventana de dos pagos ---------------------------
    escribir_rentas(wb, registros, posicion, fecha_pago, avisos, estado)

    # --- datos planos para la unificacion mensual de la CCSS ----------------
    escribir_datos_ccss(wb, registros, dias_periodo, fecha_pago)

    # --- hoja de autocomprobacion -------------------------------------------
    if "Verificacion" in wb.sheetnames:
        del wb["Verificacion"]
    totales = escribir_verificacion(wb, registros, posicion, avisos, estado)

    # Excel debe recalcular TODO al abrir el archivo, sin que nadie presione nada.
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcCompleted = False

    wb.save(salida)
    return salida, totales
    return salida


# ----------------------------------------------------------------------------
# Reporte de revision
# ----------------------------------------------------------------------------

TITULO = Font(bold=True, color="FFFFFF", name="Arial", size=11)
RELLENO = PatternFill("solid", fgColor="FF44546A")
ROJO = PatternFill("solid", fgColor="FFFFC7CE")
AMBAR = PatternFill("solid", fgColor="FFFFEB9C")


SECCIONES = [
    ("ciclo", "0 Control del ciclo bisemanal",
     ["Que se reviso", "Lo que hay", "Observacion"],
     "El pago es BISEMANAL: 14 dias exactos, 26 pagos al ano. Como los periodos "
     "no calzan con el inicio ni el fin de mes, es facil dejar un hueco o repetir "
     "dias entre un periodo y otro. Si esta hoja esta vacia, el ciclo va bien."),
    ("renta", "0b Impuesto de renta",
     ["Carne", "Que se reviso", "Base imponible", "Renta"],
     "ATENCION: los tramos del impuesto son MENSUALES pero el pago es bisemanal. "
     "Esta herramienta NO recalcula la hoja 'Rentas': usa la base que ya traia la "
     "Madre. Actualicela antes de pagar."),
    ("sin_carne", "1 Sin carne",
     ["Cedula en el archivo de horas", "Nombre", "Que hacer"],
     "Estas personas NO se pudieron cruzar con la Base de Datos del Personal y "
     "QUEDARON FUERA de la planilla generada. Agreguelas a la hoja "
     "'Base de datos del Personal' o corrija la cedula en la hoja 'Errores'."),
    ("salario", "2 Diferencia de salario",
     ["Carne", "Nombre", "Salario en horas (DCC)", "Salario en Base de Datos",
      "Diferencia", "Que hacer"],
     "La Madre SIEMPRE calcula con el salario de la Base de Datos. Si el correcto "
     "es el del archivo de horas, actualice la columna 'Salario Mensual' de la "
     "hoja 'Base de datos del Personal'."),
    ("sin_justificar", "3 Dias sin justificar",
     ["Carne", "Nombre", "Dias", "Fechas", "Tratamiento aplicado"],
     "Dias sin horas y SIN color en el archivo de horas, de personas ya "
     "contratadas. Verifique si eran ausencia, permiso sin goce o incapacidad."),
    ("jornada", "4 Jornada mixta y nocturna",
     ["Carne", "Nombre", "Horas mixtas", "Horas nocturnas",
      "Monto del diferencial", "Tratamiento aplicado"],
     "OJO: la Madre paga los 14 dias completos y ADEMAS suma un "
     "diferencial por cada hora mixta o nocturna. Como el dia mixto/nocturno ya "
     "esta pagado dentro del periodo, ese diferencial es un pago de mas. "
     "Aqui esta el monto exacto por persona."),
    ("parciales", "5 Dias con horas incompletas",
     ["Carne", "Nombre", "Equivalente en dias", "Fechas y horas trabajadas",
      "Tratamiento aplicado"],
     "Dias con menos horas que la jornada completa. Se rebajaron en proporcion "
     "para que el pago coincida con lo trabajado."),
    ("reconciliacion", "6 Reconciliacion",
     ["Carne", "Nombre", "Bruto segun archivo de horas", "Bruto que dara la Madre",
      "Diferencia", "Causa probable"],
     "Comparacion persona por persona. Una diferencia NO siempre es un error: la "
     "Madre paga los 14 dias completos y rebaja ausencias, el archivo de horas paga "
     "lo trabajado. Revise las que tengan causa marcada."),
    ("enlace", "7 Enlaces rotos en el archivo de horas",
     ["Carne", "Nombre", "Detalle", "Consecuencia"],
     "En el archivo de horas hay un dato digitado en la hoja 'Horas de trabajo' "
     "que no quedo enlazado a su hoja 'Nomina', asi que ahi no se pago. "
     "Esta planilla si lo esta pagando. Confirme cual es el correcto."),
    ("sin_horas", "8 En Madre sin horas",
     ["Carne", "Nombre", "Que hacer"],
     "Personas que estaban en la planilla anterior y no aparecen en el archivo de "
     "horas de este periodo. No se incluyeron."),

    ("nota", "9 Notas del periodo",
     ["Carne", "Nombre", "Nota"],
     "Novedades escritas en el archivo de horas (despidos, renuncias, cambios)."),
]


def hoja_reporte(wb, nombre, encabezados, filas, nota=None, resaltar=None):
    ws = wb.create_sheet(nombre[:31])
    fila = 1
    if nota:
        ws.cell(1, 1).value = nota
        ws.cell(1, 1).font = Font(italic=True, name="Arial", size=10)
        fila = 3
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(fila, j)
        c.value = h
        c.font = TITULO
        c.fill = RELLENO
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, r in enumerate(filas, start=fila + 1):
        for j, v in enumerate(r, start=1):
            c = ws.cell(i, j)
            c.value = v
            c.font = Font(name="Arial", size=10)
            if resaltar and resaltar(r):
                c.fill = AMBAR
    for j, h in enumerate(encabezados, start=1):
        ancho = max(12, min(46, len(str(h)) + 4,
                            max([len(str(h)) + 4] + [len(str(r[j - 1])) + 3
                                                     for r in filas[:200]
                                                     if j - 1 < len(r) and r[j - 1] is not None])))
        ws.column_dimensions[get_column_letter(j)].width = ancho
    ws.freeze_panes = ws.cell(fila + 1, 1)
    return ws


def escribir_reporte(ruta, avisos, resumen):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 24
    ws.cell(1, 1).value = "REPORTE DE REVISION - PLANILLA BISEMANAL"
    ws.cell(1, 1).font = Font(bold=True, size=14, name="Arial")
    f = 3
    for etiqueta, valor in resumen:
        if etiqueta == "":
            f += 1
            continue
        ws.cell(f, 1).value = etiqueta
        ws.cell(f, 1).font = Font(name="Arial", size=11,
                                  bold=etiqueta.isupper())
        ws.cell(f, 2).value = valor
        ws.cell(f, 2).font = Font(name="Arial", size=11, bold=True)
        ws.cell(f, 2).alignment = Alignment(horizontal="right")
        f += 1

    secciones = SECCIONES
    for clave, titulo, enc, nota in secciones:
        filas = avisos.items.get(clave, [])
        hoja_reporte(wb, titulo, enc, filas, nota)
    wb.save(ruta)


# ----------------------------------------------------------------------------
# Proceso principal
# ----------------------------------------------------------------------------

def detectar_periodo(ruta):
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb[H_HOJA]
    dias = []
    col = H_COL_DIA1
    while True:
        v = a_fecha(ws.cell(H_FILA_ENC, col).value)
        if v is None:
            break
        dias.append(v)
        col += 1
    wb.close()
    return dias


def procesar(ruta_horas, ruta_madre, carpeta_salida, fecha_pago,
             dias_blanco_como_ausencia=True, prorratear_ingreso=True,
             jornada_mixta="madre", log=print):
    """
    jornada_mixta:
      "madre"  -> se respeta la formula actual de la Madre. La Madre suma un
                  diferencial por cada hora mixta/nocturna ADEMAS del salario
                  del periodo, que ya cubre esos dias. Paga de mas.
      "neutro" -> las horas mixtas/nocturnas se convierten a su equivalente en
                  dias y se cargan como ordinarias, para que el pago coincida
                  con el archivo de horas. Las horas extra mixtas/nocturnas se
                  mantienen en su columna (esas si estan bien).
    """

    dias_periodo = detectar_periodo(ruta_horas)
    if not dias_periodo:
        raise RuntimeError("No se pudieron leer las fechas del periodo "
                           "(fila 2 de la hoja '%s')." % H_HOJA)
    ini, fin = dias_periodo[0], dias_periodo[-1]
    ndias = len(dias_periodo)

    log("  Periodo: %s al %s  (%d dias)" % (ini.strftime("%d/%m/%Y"),
                                            fin.strftime("%d/%m/%Y"), ndias))
    log("  Leyendo horas y colores...")
    personas = leer_horas(ruta_horas, dias_periodo)
    log("    %d personas en el archivo de horas" % len(personas))

    log("  Cruzando con la Base de Datos del Personal...")
    avisos = Aviso()
    wbm = openpyxl.load_workbook(ruta_madre, data_only=True)

    # La PLANILLA MADRE es solo plantilla: estructura, formulas y Base de Datos
    # del Personal. Nunca se modifica ni se usa para arrastrar el periodo.
    # El ESTADO (donde cerro el periodo anterior y en que mitad va la ventana de
    # renta) sale de la ultima planilla generada.
    base_carpeta = carpeta_salida
    ruta_estado, nombre_estado, primera = elegir_estado(
        base_carpeta, ruta_madre, fecha_pago)
    estado = leer_estado(ruta_estado)
    prev_ini, prev_fin, prev_pago = estado["ini"], estado["fin"], estado["pago"]
    log("    Plantilla: %s" % os.path.basename(ruta_madre))
    log("    Continua desde: %s%s" % (nombre_estado,
                                      "  (primera corrida)" if primera else ""))
    avisos.add("origen", [
        "De donde sale la estructura", os.path.basename(ruta_madre),
        "La Planilla Madre se usa como plantilla fija y NO se modifica."])
    avisos.add("origen", [
        "De donde sale el estado del ciclo", nombre_estado,
        "Primera corrida: no habia planillas generadas, se arranca de la "
        "plantilla." if primera else
        "Periodo anterior: %s al %s, pagado el %s." % (
            prev_ini.strftime("%d/%m/%Y") if prev_ini else "?",
            prev_fin.strftime("%d/%m/%Y") if prev_fin else "?",
            prev_pago.strftime("%d/%m/%Y") if prev_pago else "?")])
    correcciones = leer_correcciones(wbm)
    por_cedula, por_carne = indexar_personal(wbm)
    log("    %d correcciones de cedula tomadas de la hoja 'Errores'" % len(correcciones))

    registros = []
    usados = set()

    # --- control del ciclo bisemanal ---------------------------------------
    if ndias != DIAS_CICLO:
        avisos.add("ciclo", ["Duracion del periodo",
                             "%d dias" % ndias,
                             "El ciclo es bisemanal: deben ser %d dias exactos. "
                             "Revise la fila 2 de la hoja de horas." % DIAS_CICLO])
    if prev_fin:
        esperado = prev_fin + datetime.timedelta(days=1)
        if ini != esperado:
            hueco = (ini - esperado).days
            avisos.add("ciclo", [
                "Continuidad con el periodo anterior",
                "el anterior cerro el %s y este arranca el %s"
                % (prev_fin.strftime("%d/%m/%Y"), ini.strftime("%d/%m/%Y")),
                ("QUEDARON %d DIAS SIN PAGAR entre un periodo y otro" % hueco)
                if hueco > 0 else
                ("SE ESTAN PAGANDO %d DIAS DOS VECES" % abs(hueco))])
    if prev_pago:
        esperado_pago = prev_pago + datetime.timedelta(days=DIAS_CICLO)
        if fecha_pago != esperado_pago:
            avisos.add("ciclo", [
                "Fecha de pago", fecha_pago.strftime("%d/%m/%Y"),
                "Con cadencia bisemanal correspondia el %s"
                % esperado_pago.strftime("%d/%m/%Y")])

    pagos_mes = pagos_del_mes(fecha_pago)
    if len(pagos_mes) >= 3:
        avisos.add("ciclo", [
            "Mes con 3 pagos",
            "%s: %s" % (fecha_pago.strftime("%m/%Y"),
                        ", ".join(d.strftime("%d/%m") for d in pagos_mes)),
            "Aplica el corrimiento: el tercer pago no entra en la base de renta "
            "de este mes, arranca la ventana siguiente."])

    for p in personas:
        clasificar_dias(p, dias_periodo, dias_blanco_como_ausencia, prorratear_ingreso)

        ced = p.cedula
        reg = por_cedula.get(ced)
        if reg is None and ced in correcciones:
            reg = por_cedula.get(correcciones[ced])
        if reg is None:
            avisos.add("sin_carne", [p.cedula, p.nombre,
                                     "Agregar a la Base de Datos del Personal"])
            continue

        registros.append({"carne": reg["carne"], "persona": p, "bd": reg})
        usados.add(reg["carne"])

        # --- diferencial de jornada mixta / nocturna -------------------------
        p.horas_mixtas_orig, p.horas_noct_orig = p.horas[7], p.horas[6]
        s_bd0 = reg["salario"] if isinstance(reg["salario"], (int, float)) \
            else (p.salario_dcc or 0)
        if p.horas[7] or p.horas[6]:
            dif_j = (p.horas[7] * (s_bd0 / 30.0 / 7 - s_bd0 / 30.0 / 8)
                     + p.horas[6] * (s_bd0 / 30.0 / 6 - s_bd0 / 30.0 / 8))
            avisos.add("jornada", [
                reg["carne"], p.nombre, round(p.horas[7], 2), round(p.horas[6], 2),
                round(dif_j, 2),
                "Se pago de mas (formula actual de la Madre)" if jornada_mixta == "madre"
                else "Neutralizado: se cargo como horas ordinarias equivalentes"])
            if jornada_mixta == "neutro":
                p.horas[8] += p.horas[7] / 7.0 * 8 + p.horas[6] / 6.0 * 8
                p.horas[7] = 0.0
                p.horas[6] = 0.0

        # --- avisos ---------------------------------------------------------
        s_bd, s_dcc = reg["salario"], p.salario_dcc
        if isinstance(s_bd, (int, float)) and isinstance(s_dcc, (int, float)):
            if abs(s_bd - s_dcc) > 1:
                avisos.add("salario", [reg["carne"], p.nombre, round(s_dcc, 2),
                                       round(s_bd, 2), round(s_dcc - s_bd, 2),
                                       "Revisar cual es el vigente"])
        elif not isinstance(s_bd, (int, float)):
            avisos.add("salario", [reg["carne"], p.nombre, s_dcc, "(vacio)", "",
                                   "Falta el salario en la Base de Datos"])

        if p.d_sin_justificar:
            fechas = ", ".join(f.strftime("%d/%m") for f, _ in p.d_sin_justificar[:12])
            if len(p.d_sin_justificar) > 12:
                fechas += " ..."
            avisos.add("sin_justificar", [
                reg["carne"], p.nombre,
                round(sum(x for _, x in p.d_sin_justificar), 2), fechas,
                "Cargado como AUSENCIA" if dias_blanco_como_ausencia
                else "NO se rebajo (se paga completo)"])

        if p.d_parciales:
            fechas = ", ".join("%s (%s h)" % (f.strftime("%d/%m"), round(8 * (1 - x), 2))
                               for f, x in p.d_parciales[:8])
            if len(p.d_parciales) > 8:
                fechas += " ..."
            avisos.add("parciales", [
                reg["carne"], p.nombre,
                round(sum(x for _, x in p.d_parciales), 2), fechas,
                "Rebajado en proporcion"])

        if p.novedades:
            avisos.add("nota", [reg["carne"], p.nombre, " / ".join(p.novedades)])

        if p.enlaces_rotos:
            avisos.add("enlace", [reg["carne"], p.nombre, "; ".join(p.enlaces_rotos),
                                  "Ese dato NO se pago en el archivo de horas. "
                                  "La Madre si lo esta pagando."])

        # --- reconciliacion: replica el calculo de la Madre ------------------
        h7, h6 = p.horas_mixtas_orig, p.horas_noct_orig
        dias_equiv = p.horas[8] / 8.0 + p.horas[7] / 7.0 + p.horas[6] / 6.0
        sal = s_bd if isinstance(s_bd, (int, float)) else (s_dcc or 0)
        sal_dia = sal / 30.0
        sal_hora = sal / 30.0 / 8.0
        h_mixta = sal / 30.0 / 7.0
        h_noct = sal / 30.0 / 6.0
        bruto_madre = (
            sal_dia * ndias
            + p.extras[8] * sal_hora * 1.5
            + p.dias_dobles * sal_dia
            + p.extra_dobles * sal_hora * 2 * 1.5
            + p.horas[7] * (h_mixta - sal_hora)
            + p.extras[7] * h_mixta * 1.5
            + p.horas[6] * (h_noct - sal_hora)
            + p.extras[6] * h_noct * 1.5
            - p.d_ausencia * sal_dia
            - p.d_psg * sal_dia
            - p.d_ccss * sal_dia * 0.5
            - p.d_ccss_mas3 * sal_dia
            - p.d_ins * sal_dia
            + p.incentivos + p.ajuste_monto
        )
        p.bruto_madre = bruto_madre
        dif = bruto_madre - p.bruto_dcc
        if abs(dif) > 1:
            causas = []
            if isinstance(s_bd, (int, float)) and isinstance(s_dcc, (int, float)) \
                    and abs(s_bd - s_dcc) > 1:
                causas.append("salario distinto entre los dos archivos")
            if jornada_mixta == "madre" and (p.horas[7] or p.horas[6]):
                causas.append("diferencial de jornada mixta/nocturna")
            if p.enlaces_rotos:
                causas.append("al archivo de horas le falta un enlace "
                              "(NO se pago ahi, la Madre si lo paga)")
            if p.extra_dobles and (h7 or h6):
                causas.append("tarifa de extras dobles: la Madre usa la tarifa "
                              "diurna y el archivo de horas la de la jornada")
            if dias_equiv > ndias + 0.01:
                causas.append("registro %.2f dias-equivalentes en un periodo de %d "
                              "(revisar horas del dia)" % (dias_equiv, ndias))
            if p.d_sin_justificar:
                causas.append("dias sin justificar")
            if p.ingreso and p.ingreso > ini:
                causas.append("ingreso a mitad de periodo")
            avisos.add("reconciliacion", [
                reg["carne"], p.nombre, round(p.bruto_dcc, 2), round(bruto_madre, 2),
                round(dif, 2), ", ".join(causas) or "revisar"])

    for carne, reg in por_carne.items():
        pass  # el roster anterior se compara abajo

    # personas que estaban en la planilla anterior y ya no vienen
    for carne in sorted(estado["roster"]):
        if carne not in usados and carne in por_carne:
            avisos.add("sin_horas", [carne, "", "No vino en el archivo de horas"])

    # --- impuesto de renta: ventana de DOS pagos con corrimiento ------------
    etiq_1, etiq_2 = estado["etiquetas"]
    prim, segu = estado["prim"], estado["segu"]
    posicion = estado["posicion"]
    tramos_v = tramos_de_renta(openpyxl.load_workbook(
        ruta_estado, data_only=True)[M_RENTAS])
    con_renta = []
    for c, slots in estado["rentas"].items():
        b = num(slots.get("D")) + num(slots.get("I"))
        r = renta_de(b, tramos_v)
        if r > 0:
            con_renta.append((c, b, r))

    avisos.add("renta", ["(ventana)", "Posicion de este pago",
                         "PRIMER pago de la ventana" if posicion == 1
                         else "SEGUNDO pago de la ventana",
                         ("La ventana queda abierta: la renta se retiene hasta el "
                          "proximo pago (%s)." %
                          (fecha_pago + datetime.timedelta(days=DIAS_CICLO)
                           ).strftime("%d/%m/%Y")) if posicion == 1 else
                         "La ventana se cierra con este pago: aqui se retiene la "
                         "renta de los dos pagos."])
    avisos.add("renta", ["(ventana)", "Casillas de la hoja Rentas",
                         "%s (%d con dato)  |  %s (%d con dato)"
                         % (etiq_1, prim, etiq_2, segu),
                         "Los rotulos van corridos respecto al mes calendario: "
                         "es lo esperado con el corrimiento."])
    avisos.add("renta", ["(ventana)", "Ventanas por ano",
                         "13 (26 pagos / 2)",
                         "Los tramos del impuesto son MENSUALES. Con 13 ventanas se "
                         "aplica 13 veces el minimo exento de 918.000 en lugar de 12, "
                         "y cada base son %d dias en vez de 30. Confirmelo con "
                         "contabilidad." % (2 * DIAS_CICLO)])
    avisos.add("renta", ["(ventana)", "Valores que trae la Madre",
                         "%d personas con renta calculada" % len(con_renta),
                         ("Este pago ABRE la ventana: la herramienta cargo el bruto en la "
                          "1a casilla, dejo la 2a vacia y puso la retencion en CERO. "
                          "La renta se cobra completa en el pago del %s."
                          % (fecha_pago + datetime.timedelta(days=DIAS_CICLO)
                             ).strftime("%d/%m/%Y"))
                         if posicion == 1 else
                         ("Este pago CIERRA la ventana: la herramienta cargo el bruto "
                          "en la 2a casilla y aqui se retiene la renta de los dos "
                          "pagos.")])
    for c, b, r in sorted(con_renta, key=lambda x: -x[2])[:200]:
        avisos.add("renta", [c, "", round(b, 2), round(r, 2)])

    # --- escritura ----------------------------------------------------------
    etiqueta = fecha_pago.strftime("%d-%m-%Y")
    destino = carpeta_generadas(base_carpeta)
    salida = os.path.join(destino, "%s%s.xlsx" % (PREFIJO_SALIDA, etiqueta))
    reporte = os.path.join(destino, "%s%s.xlsx" % (PREFIJO_REPORTE, etiqueta))

    log("  Escribiendo la Planilla Madre...")
    _, totales = escribir_madre(ruta_madre, salida, registros, dias_periodo,
                                fecha_pago, avisos, posicion, estado)

    tot = lambda k: sum(getattr(r["persona"], k) for r in registros)
    resumen = [
        ("Periodo (bisemanal, %d dias)" % ndias, "%s al %s" % (ini.strftime("%d/%m/%Y"), fin.strftime("%d/%m/%Y"))),
        ("Fecha de pago", fecha_pago.strftime("%d/%m/%Y")),
        ("Proximo periodo",
         "%s al %s" % ((fin + datetime.timedelta(days=1)).strftime("%d/%m/%Y"),
                       (fin + datetime.timedelta(days=DIAS_CICLO)).strftime("%d/%m/%Y"))),
        ("Proxima fecha de pago",
         (fecha_pago + datetime.timedelta(days=DIAS_CICLO)).strftime("%d/%m/%Y")),
        ("Pagos que caen en este mes", len(pagos_mes)),
        ("", ""),
        ("ARCHIVOS", ""),
        ("Plantilla usada (no se modifica)", os.path.basename(ruta_madre)),
        ("Continua desde", nombre_estado),
        ("", ""),
        ("TRABAJADORES", ""),
        ("En el archivo de horas", len(personas)),
        ("Incluidos en la planilla", len(registros)),
        ("Sin carne (quedaron fuera)", avisos.total("sin_carne")),
        ("En la planilla anterior y sin horas ahora", avisos.total("sin_horas")),
        ("", ""),
        ("HORAS", ""),
        ("Ordinarias (jornada diurna)", round(sum(r["persona"].horas[8] for r in registros), 2)),
        ("Mixtas", round(sum(r["persona"].horas[7] for r in registros), 2)),
        ("Nocturnas", round(sum(r["persona"].horas[6] for r in registros), 2)),
        ("Extras 150%", round(sum(r["persona"].extras[8] for r in registros), 2)),
        ("Extras dobles 200%", round(sum(r["persona"].extra_dobles for r in registros), 2)),
        ("Dias dobles", round(sum(r["persona"].dias_dobles for r in registros), 2)),
        ("", ""),
        ("DIAS NO TRABAJADOS (leidos de los colores)", ""),
        ("Ausencia", round(tot("d_ausencia"), 2)),
        ("  de los cuales, por ingreso a mitad de periodo", round(tot("d_previos"), 2)),
        ("Permiso sin goce (PSG)", round(tot("d_psg"), 2)),
        ("Incapacidad CCSS primeros 3 dias", round(tot("d_ccss"), 2)),
        ("Incapacidad CCSS mas de 3 dias", round(tot("d_ccss_mas3"), 2)),
        ("Incapacidad INS", round(tot("d_ins"), 2)),
        ("", ""),
        ("TOTALES DEL PAGO", ""),
        ("Salario bruto", round(totales["bruto"], 2)),
        ("Rebajo CCSS 10.83%", round(totales["ccss"], 2)),
        ("Retencion de renta", round(totales["renta"], 2)),
        ("Salario neto a pagar", round(totales["neto"], 2)),
        ("", ""),
        ("PENDIENTES DE REVISAR", ""),
        ("Avisos del ciclo bisemanal", avisos.total("ciclo")),
        ("Diferencias de salario", avisos.total("salario")),
        ("Personas con dias sin justificar", avisos.total("sin_justificar")),
        ("Personas con dias de horas incompletas", avisos.total("parciales")),
        ("Personas con jornada mixta/nocturna", avisos.total("jornada")),
        ("  diferencial que suma la Madre (colones)",
         round(sum(f[4] for f in avisos.items.get("jornada", [])), 2)),
        ("  tratamiento", "se dejo como esta la Madre"
         if jornada_mixta == "madre" else "neutralizado"),
        ("Enlaces rotos en el archivo de horas", avisos.total("enlace")),
        ("Diferencias de bruto vs archivo de horas", avisos.total("reconciliacion")),
    ]

    log("  Escribiendo el reporte de revision...")
    escribir_reporte(reporte, avisos, resumen)
    return salida, reporte, resumen, avisos


# ----------------------------------------------------------------------------
# Asistente interactivo
# ----------------------------------------------------------------------------

def buscar(carpeta, incluye, excluye=()):
    cand = []
    for n in os.listdir(carpeta):
        if not n.lower().endswith((".xlsx", ".xlsm")) or n.startswith("~$"):
            continue
        bajo = sin_tildes(n)
        if all(t in bajo for t in incluye) and not any(t in bajo for t in excluye):
            cand.append(os.path.join(carpeta, n))
    cand.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return cand


def preguntar_si(texto, defecto=True):
    suf = "[ENTER = SI, escriba n para NO]" if defecto else "[ENTER = NO, escriba s para SI]"
    r = input("  %s %s: " % (texto, suf)).strip().lower()
    if not r:
        return defecto
    return r[0] in ("s", "y", "1")


def elegir(candidatos, titulo):
    if not candidatos:
        return None
    if len(candidatos) == 1:
        return candidatos[0]
    print("\n  Hay varios archivos de %s. Elija uno:" % titulo)
    for i, c in enumerate(candidatos, 1):
        print("    %d) %s" % (i, os.path.basename(c)))
    while True:
        r = input("  Numero [1]: ").strip() or "1"
        if r.isdigit() and 1 <= int(r) <= len(candidatos):
            return candidatos[int(r) - 1]


def pedir_fecha(texto, defecto):
    while True:
        r = input("  %s [ENTER = %s]: " % (texto, defecto.strftime("%d/%m/%Y"))).strip()
        if not r:
            return defecto
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(r, fmt).date()
            except ValueError:
                pass
        print("     Formato no valido. Use dia/mes/año, por ejemplo 28/08/2026")


def interactivo(carpeta):
    print("")
    print("  ==========================================================")
    print("     GENERADOR DE PLANILLA BISEMANAL")
    print("     Ingenieria Estrella S.A.")
    print("  ==========================================================")
    print("")
    print("  Carpeta: %s" % carpeta)
    print("")

    horas = elegir(buscar(carpeta, ["trabajadores"], ["madre", "lista", "reporte"])
                   or buscar(carpeta, ["dcc"], ["madre", "lista", "reporte"]),
                   "horas")
    madre = elegir(buscar(carpeta, ["madre"], ["lista", "reporte"]), "planilla madre")

    if not horas or not madre:
        print("  NO ENCONTRE LOS ARCHIVOS.")
        print("  Coloque en esta carpeta el archivo de horas (el que dice")
        print("  'Trabajadores') y la 'Planilla Madre', y vuelva a intentar.")
        return 1

    dias = detectar_periodo(horas)
    print("  Archivo de HORAS:")
    print("     %s" % os.path.basename(horas))
    if dias:
        print("     Periodo detectado: %s al %s (%d dias)"
              % (dias[0].strftime("%d/%m/%Y"), dias[-1].strftime("%d/%m/%Y"), len(dias)))
    print("")
    print("  PLANILLA MADRE:")
    print("     %s" % os.path.basename(madre))
    print("")

    if not preguntar_si("Son estos los archivos correctos?"):
        print("  Cancelado. Deje solo los archivos correctos en la carpeta.")
        return 1

    print("")
    pago = pedir_fecha("Fecha de PAGO",
                       (dias[-1] + datetime.timedelta(days=12)) if dias
                       else datetime.date.today())
    print("")
    print("  Dos preguntas mas (si no esta seguro, presione ENTER):")
    print("")
    blanco = preguntar_si("Los dias en blanco SIN color se rebajan como ausencia?")
    prorr = preguntar_si("A quien entro a mitad del periodo se le paga solo desde su ingreso?")
    print("")
    print("  La Madre suma un extra por cada hora de jornada mixta o nocturna,")
    print("  ademas del salario quincenal que ya cubre esos dias (paga de mas).")
    neutro = preguntar_si("Quiere corregir eso automaticamente?", defecto=False)

    print("")
    print("  ----------------------------------------------------------")
    print("  Procesando...")
    try:
        salida, reporte, resumen, avisos = procesar(
            horas, madre, carpeta, pago,
            dias_blanco_como_ausencia=blanco,
            prorratear_ingreso=prorr,
            jornada_mixta=("neutro" if neutro else "madre"),
            log=print)
    except Exception as e:
        print("")
        print("  OCURRIO UN ERROR: %s" % e)
        print("")
        import traceback
        traceback.print_exc()
        input("  Presione ENTER para cerrar.")
        return 1

    print("  ----------------------------------------------------------")
    print("")
    print("  LISTO. Se crearon 2 archivos en esta carpeta:")
    print("")
    print("    1) %s" % os.path.basename(salida))
    print("       <- esta es la planilla. Abrala en Excel.")
    print("    2) %s" % os.path.basename(reporte))
    print("       <- REVISE ESTE PRIMERO. Trae lo que quedo pendiente.")
    print("")
    print("  Resumen:")
    for etiqueta, valor in resumen:
        if etiqueta == "":
            print("")
        elif valor == "":
            print("    %s" % etiqueta)
        else:
            print("      %-46s %s" % (etiqueta, valor))
    print("")
    print("  IMPORTANTE: al abrir la planilla en Excel, presione Ctrl+Alt+F9")
    print("  (o Cmd+Alt+F9 en Mac) para que recalcule todas las formulas.")
    print("")
    try:
        os.system('open -R "%s"' % salida)
    except Exception:
        pass
    input("  Presione ENTER para cerrar esta ventana.")
    return 0


def main():
    args = sys.argv[1:]
    carpeta = os.path.dirname(os.path.abspath(__file__))
    carpeta = os.path.dirname(carpeta) if os.path.basename(carpeta) == "programa" else carpeta

    if not args or args[0] == "--interactivo":
        if len(args) > 1:
            carpeta = args[1]
        return interactivo(carpeta)

    # modo linea de comandos
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("horas")
    ap.add_argument("--madre", required=True)
    ap.add_argument("--pago", required=True)
    ap.add_argument("--salida", default=None)
    ap.add_argument("--blancos", choices=["ausencia", "ignorar"], default="ausencia")
    ap.add_argument("--ingresos", choices=["prorratear", "completo"], default="prorratear")
    ap.add_argument("--mixtas", choices=["madre", "neutro"], default="madre")
    a = ap.parse_args(args)
    pago = datetime.datetime.strptime(a.pago, "%Y-%m-%d").date()
    salida, reporte, resumen, _ = procesar(
        a.horas, a.madre, a.salida or os.path.dirname(os.path.abspath(a.horas)), pago,
        dias_blanco_como_ausencia=(a.blancos == "ausencia"),
        prorratear_ingreso=(a.ingresos == "prorratear"),
        jornada_mixta=a.mixtas)
    print("\nPlanilla: %s\nReporte : %s\n" % (salida, reporte))
    for k, v in resumen:
        if k and v != "":
            print("  %-46s %s" % (k, v))
    return 0


if __name__ == "__main__":
    sys.exit(main())


# ----------------------------------------------------------------------------
# Unificacion mensual para la CCSS
# ----------------------------------------------------------------------------

UF_COLUMNAS = ["Carné", "Días Laborados", "Puesto", "Nombre del funcionario",
               "Identificación", "Total Salario Bruto", "Estado", "Fecha IC",
               "Fecha EX", "Reportado de manera anticipada", "Salario a Reportar"]


def leer_datos_ccss(ruta):
    """Lee la hoja de valores planos que deja cada planilla generada."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    if HOJA_CCSS not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            "El archivo '%s' no trae la hoja '%s'. Fue generado con una version "
            "anterior del programa: vuelva a generarlo." %
            (os.path.basename(ruta), HOJA_CCSS))
    ws = wb[HOJA_CCSS]
    per = {"archivo": os.path.basename(ruta),
           "ini": a_fecha(ws["B2"].value),
           "fin": a_fecha(ws["D2"].value),
           "pago": a_fecha(ws["F2"].value),
           "gente": []}
    for f in range(5, ws.max_row + 1):
        car = ws.cell(f, 1).value
        if not car:
            continue
        per["gente"].append({
            "carne": str(car).strip(),
            "cedula": ws.cell(f, 2).value,
            "nombre": ws.cell(f, 3).value,
            "puesto": ws.cell(f, 4).value,
            "dias": num(ws.cell(f, 5).value),
            "bruto": num(ws.cell(f, 6).value),
            "ins": num(ws.cell(f, 7).value),
            "ingreso": a_fecha(ws.cell(f, 8).value),
            "novedad": ws.cell(f, 9).value or "",
            "origen_ing": ws.cell(f, 10).value or "",
        })
    wb.close()
    return per


def hay_salida(texto):
    t = sin_tildes(texto)
    return any(p in t for p in PALABRAS_SALIDA)


def fecha_en_texto(texto):
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", str(texto or ""))
    if not m:
        return None
    d, mes, a = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if a < 100:
        a += 2000
    try:
        return datetime.date(a, mes, d)
    except ValueError:
        return None


def unificar(rutas, carpeta_salida, etiqueta=None, log=print):
    """Une N planillas generadas en el formato de reporte para la CCSS."""
    if not rutas:
        raise RuntimeError("No se escogio ningun archivo.")

    periodos = [leer_datos_ccss(r) for r in rutas]
    periodos.sort(key=lambda p: p["pago"] or datetime.date(1900, 1, 1))
    ini = min(p["ini"] for p in periodos if p["ini"])
    fin = max(p["fin"] for p in periodos if p["fin"])
    log("  Periodos: %d  (%s al %s)" % (len(periodos),
                                        ini.strftime("%d/%m/%Y"),
                                        fin.strftime("%d/%m/%Y")))

    gente = {}
    for per in periodos:
        for g in per["gente"]:
            r = gente.setdefault(g["carne"], {
                "carne": g["carne"], "cedula": g["cedula"], "nombre": g["nombre"],
                "puesto": g["puesto"], "dias": 0.0, "bruto": 0.0, "ins": 0.0,
                "ingreso": g["ingreso"], "origen_ing": g["origen_ing"],
                "novedades": [], "periodos": [],
            })
            r["dias"] += g["dias"]
            r["bruto"] += g["bruto"]
            r["ins"] += g["ins"]
            # los datos personales se toman del periodo mas reciente
            for k in ("cedula", "nombre", "puesto", "ingreso", "origen_ing"):
                if g[k]:
                    r[k] = g[k]
            if g["novedad"]:
                r["novedades"].append(g["novedad"])
            r["periodos"].append((per["pago"], g["dias"], g["bruto"]))

    avisos = Aviso()
    filas = []
    for car in sorted(gente, key=lambda c: sin_tildes(gente[c]["nombre"] or c)):
        r = gente[car]
        nota = " / ".join(r["novedades"])
        salio = hay_salida(nota)
        entro = bool(r["ingreso"] and ini <= r["ingreso"] <= fin)
        con_ins = r["ins"] > 0

        if salio:
            estado = "EX"
        elif entro:
            estado = "IC"
        elif con_ins:
            estado = "INS"
        else:
            estado = "SAL"

        f_ic = r["ingreso"] if entro else None
        f_ex = fecha_en_texto(nota) if salio else None

        filas.append({
            "carne": car, "dias": round(r["dias"], 2), "puesto": r["puesto"],
            "nombre": r["nombre"], "cedula": r["cedula"],
            "bruto": round(r["bruto"], 2), "estado": estado,
            "f_ic": f_ic, "f_ex": f_ex, "nota": nota,
            "periodos": len(r["periodos"]),
        })

        motivos = [m for m, ok in (("salida", salio), ("ingreso", entro),
                                   ("incapacidad INS", con_ins)) if ok]
        if len(motivos) > 1:
            avisos.add("multiple", [car, r["nombre"], estado,
                                    "Cumple mas de una condicion: " + ", ".join(motivos),
                                    nota])
        if salio and not f_ex:
            avisos.add("sin_fecha", [car, r["nombre"], "EX",
                                     "Se marco salida pero la nota no trae fecha",
                                     nota])
        if entro and not (r.get("origen_ing") or "").startswith("Base de Datos"):
            avisos.add("ingreso_dudoso", [
                car, r["nombre"],
                r["ingreso"].strftime("%d/%m/%Y") if r["ingreso"] else "",
                r.get("origen_ing") or "",
                "Agregue la fecha de ingreso en la Base de Datos del Personal"])
        if estado != "SAL":
            avisos.add("movimiento", [car, r["nombre"], estado,
                                      f_ic.strftime("%d/%m/%Y") if f_ic else "",
                                      f_ex.strftime("%d/%m/%Y") if f_ex else "",
                                      nota])
        if len(r["periodos"]) < len(periodos):
            faltan = [p["pago"].strftime("%d/%m/%Y") for p in periodos
                      if p["pago"] not in [x[0] for x in r["periodos"]]]
            avisos.add("parcial", [car, r["nombre"], len(r["periodos"]),
                                   len(periodos), ", ".join(faltan)])

    etiqueta = etiqueta or "%s_%s" % (ini.strftime("%d-%m-%Y"), fin.strftime("%d-%m-%Y"))
    destino = carpeta_generadas(carpeta_salida)
    salida = os.path.join(destino, "Unificacion CCSS %s.xlsx" % etiqueta)
    escribir_unificacion(salida, filas, periodos, avisos, ini, fin)
    return salida, filas, periodos, avisos


def escribir_unificacion(ruta, filas, periodos, avisos, ini, fin):
    wb = openpyxl.Workbook()

    # ---- Union Final -------------------------------------------------------
    ws = wb.active
    ws.title = "Union Final"
    for j, t in enumerate(UF_COLUMNAS, start=1):
        c = ws.cell(1, j)
        c.value = t
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", fgColor="FF44546A")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, f in enumerate(filas, start=2):
        ws.cell(i, 1).value = f["carne"]
        ws.cell(i, 2).value = f["dias"]
        ws.cell(i, 3).value = f["puesto"]
        ws.cell(i, 4).value = f["nombre"]
        ws.cell(i, 5).value = f["cedula"]
        ws.cell(i, 6).value = f["bruto"]
        ws.cell(i, 7).value = f["estado"]
        ws.cell(i, 8).value = f["f_ic"]
        ws.cell(i, 9).value = f["f_ex"]
        ws.cell(i, 11).value = "=+F%d-J%d" % (i, i)
        ws.cell(i, 6).number_format = '#,##0.00'
        ws.cell(i, 11).number_format = '#,##0.00'
        for col in (8, 9):
            ws.cell(i, col).number_format = 'DD/MM/YYYY'
    ult = len(filas) + 1
    t = ult + 1
    ws.cell(t, 4).value = "TOTALES"
    ws.cell(t, 4).font = Font(bold=True, name="Arial")
    for col in (6, 10, 11):
        c = ws.cell(t, col)
        c.value = "=SUM(%s2:%s%d)" % (get_column_letter(col), get_column_letter(col), ult)
        c.font = Font(bold=True, name="Arial")
        c.number_format = '#,##0.00'
    for j, w in enumerate([10, 13, 26, 34, 16, 18, 9, 12, 12, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    # ---- Detalle por periodo ----------------------------------------------
    ws = wb.create_sheet("Detalle por periodo")
    enc = ["Carné", "Nombre", "Identificación"]
    for p in periodos:
        enc += ["Días %s" % p["pago"].strftime("%d/%m"),
                "Bruto %s" % p["pago"].strftime("%d/%m")]
    enc += ["Días total", "Bruto total"]
    for j, x in enumerate(enc, start=1):
        c = ws.cell(1, j)
        c.value = x
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", fgColor="FF44546A")
    porcarne = {}
    for p in periodos:
        for g in p["gente"]:
            porcarne.setdefault(g["carne"], {})[p["pago"]] = g
    for i, f in enumerate(filas, start=2):
        ws.cell(i, 1).value = f["carne"]
        ws.cell(i, 2).value = f["nombre"]
        ws.cell(i, 3).value = f["cedula"]
        col = 4
        for p in periodos:
            g = porcarne.get(f["carne"], {}).get(p["pago"])
            ws.cell(i, col).value = round(g["dias"], 2) if g else None
            ws.cell(i, col + 1).value = round(g["bruto"], 2) if g else None
            col += 2
        ws.cell(i, col).value = f["dias"]
        ws.cell(i, col + 1).value = f["bruto"]
    ws.freeze_panes = "D2"
    for j in range(1, len(enc) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15 if j > 3 else 26

    # ---- Revision ----------------------------------------------------------
    secciones = [
        ("movimiento", "Movimientos detectados",
         ["Carné", "Nombre", "Estado", "Fecha IC", "Fecha EX", "De donde salio"],
         "Todo lo que NO quedo como SAL. Revise que el estado y las fechas sean "
         "correctos antes de enviar a la CCSS."),
        ("sin_fecha", "Salidas sin fecha",
         ["Carné", "Nombre", "Estado", "Detalle", "Nota"],
         "Se detecto una salida pero la nota del archivo de horas no traia fecha. "
         "Hay que escribir la Fecha EX a mano."),
        ("multiple", "Casos con mas de una condicion",
         ["Carné", "Nombre", "Estado asignado", "Detalle", "Nota"],
         "Personas que cumplen mas de una condicion a la vez (por ejemplo entraron "
         "y salieron en el mismo mes). Se asigno un solo estado: confirme cual "
         "corresponde reportar."),
        ("ingreso_dudoso", "Fechas de ingreso a confirmar",
         ["Carné", "Nombre", "Fecha usada", "De donde salio", "Que hacer"],
         "Se reportan como INCLUSION (IC) pero la fecha de ingreso no estaba en la "
         "Base de Datos del Personal: se tomo del archivo de horas. Confirme la "
         "fecha antes de enviar a la CCSS y complete la Base de Datos."),
        ("parcial", "Aparecen en algunos periodos",
         ["Carné", "Nombre", "En cuantos periodos", "Periodos unidos", "Falto en"],
         "Gente que no vino en todos los periodos escogidos. Es normal en ingresos "
         "y salidas, pero conviene revisarlo."),
    ]
    for clave, titulo, encab, nota in secciones:
        hoja_reporte(wb, titulo, encab, avisos.items.get(clave, []), nota)

    # ---- Resumen -----------------------------------------------------------
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26
    ws["A1"] = "UNIFICACION PARA LA CCSS"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    est = defaultdict(int)
    for f in filas:
        est[f["estado"]] += 1
    datos = [("Periodo cubierto", "%s al %s" % (ini.strftime("%d/%m/%Y"),
                                                fin.strftime("%d/%m/%Y"))),
             ("Planillas unidas", len(periodos)), ("", "")]
    for p in periodos:
        datos.append(("   pago %s" % p["pago"].strftime("%d/%m/%Y"), p["archivo"]))
    datos += [("", ""), ("PERSONAS", ""), ("Total en el reporte", len(filas))]
    for e, n in sorted(est.items()):
        datos.append(("   %s" % {"SAL": "SAL (salario)", "IC": "IC (inclusion)",
                                 "EX": "EX (exclusion)",
                                 "INS": "INS (incapacidad)"}.get(e, e), n))
    datos += [("", ""), ("MONTOS", ""),
              ("Total salario bruto", round(sum(f["bruto"] for f in filas), 2)),
              ("Total dias laborados", round(sum(f["dias"] for f in filas), 2)),
              ("", ""), ("PENDIENTES DE REVISAR", ""),
              ("Salidas sin fecha", avisos.total("sin_fecha")),
              ("Fechas de ingreso a confirmar", avisos.total("ingreso_dudoso")),
              ("Con mas de una condicion", avisos.total("multiple")),
              ("En solo algunos periodos", avisos.total("parcial"))]
    f = 3
    for k, v2 in datos:
        if k == "":
            f += 1
            continue
        ws.cell(f, 1).value = k
        ws.cell(f, 1).font = Font(name="Arial", size=11, bold=k.isupper())
        ws.cell(f, 2).value = v2
        ws.cell(f, 2).font = Font(name="Arial", size=11, bold=True)
        ws.cell(f, 2).alignment = Alignment(horizontal="right")
        f += 1

    wb.calculation.fullCalcOnLoad = True
    wb.save(ruta)
    return ruta
